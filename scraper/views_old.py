import csv
import glob
import json
import os
import random
import threading
import time

from django.http import HttpResponse, Http404
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .radius import *
from .utils import create_text_file


# Create your views here.

def faq(request):
    return render(request, 'scraper/faq.html')


def radius_check(request):
    if request.method == 'POST':
        city_name_r = request.POST.get('radius_value')
        city_list, dist_list = get_cities(city_name_r)
        city_list = zip(city_list, dist_list)
        context = {
            'city_list': city_list,
            'city_name_r': city_name_r,
        }
        return render(request, 'scraper/radius.html', context)

    return render(request, 'scraper/radius.html')


def show(request):
    if request.method == 'POST':
        file_name_r = request.POST.get('filename')
        file_path = os.path.join(file_name_r)
        if os.path.exists(file_path):

            if request.POST.get('delete_file'):
                # deleting the file
                os.remove(file_path)
                file_names = []
                for file in glob.glob("*.xlsx"):
                    file_names.append(file)
                return render(request, 'scraper/show.html', {"file_names": file_names})
            else:
                # downloading the file
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + json.dumps(os.path.basename(file_path))
                    return response

        raise Http404

    else:
        file_names = []
        for file in glob.glob("*.xlsx"):
            file_names.append(file)
        return render(request, 'scraper/show.html', {"file_names": file_names})


@csrf_exempt
def row_ajax(request):
    if request.method == 'POST':
        return HttpResponse(row_count)


pause_thread = False


def stop_scrap(request):
    if request.method == 'POST':
        if request.POST.get('stop_scrap'):
            stop_scraping()
            if t1.is_alive():
                print("thread still alive man. Fuck")
                return render(request, 'scraper/index.html')
        else:
            global pause_value
            pause_value = pause_scraping()  # stop the scraper

            if os.path.isfile('entry.txt'):
                f = open("entry.txt", "r")
                contents = f.read()
                context = {
                    "row_count": row_count,
                    "entry": contents,
                    "running": "True",
                    "pause_scrap": pause_value,
                }
                return render(request, 'scraper/index.html', context)

        return render(request, 'scraper/index.html')


def index(request):
    if request.method == 'POST':

        global entry_r
        global choice_r
        hashtag_r = request.POST.get('hashtag')
        location_r = request.POST.get('location')
        zip_r = request.POST.get('zip')
        if hashtag_r != "":
            choice_r = "tag"
            entry_r = hashtag_r
        if zip_r != "":
            choice_r = "zip"
            entry_r = zip_r
        if location_r != "":
            choice_r = "location"
            entry_r = location_r

        if request.POST.get('startscraping'):
            global row_count
            row_count = 0
            create_text_file(entry_r)

            global t1

            t1 = threading.Thread(target=start_scraping, args=(entry_r, choice_r))
            t1.daemon = True
            t1.start()

            if entry_r != "":
                print(row_count)
                context = {
                    "row_count": row_count,
                    "entry": entry_r,
                    "running": "True",
                }

                return render(request, 'scraper/index.html', context)

        elif request.POST.get('checklocation'):
            if choice_r != "tag":
                location_list = get_location_list(entry_r, choice_r)
                context = {
                    "location_list": location_list,
                    "entry": entry_r,
                }
                return render(request, 'scraper/index.html', context)

        else:
            pass

    else:
        if os.path.isfile('entry.txt'):
            f = open("entry.txt", "r")
            contents = f.read()
            context = {
                "row_count": row_count,
                "entry": contents,
                "running": "True",
                "pause_scrap": pause_thread,
            }
            return render(request, 'scraper/index.html', context)
        else:

            return render(request, 'scraper/index.html')


num_of_pages = 500000
row_count = 0
stop_thread = False

save_data = []
user_agent = "Mozilla/5.0 (iPhone; CPU iPhone OS 12_3_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 Instagram 105.0.0.11.118 (iPhone11,8; iOS 12_3_1; en_US; en-US; scale=2.00; 828x1792; 165586599)"

cookie_value = [

    """mid=XVUXTwALAAE_gRSiZ7oReErEEsM-; csrftoken=NVdeQTMMpNZjD9c5fsBkmFmRwjwJNZkv; ds_user_id=18070511049; sessionid=18070511049%3AxUebJzsAmvkqC1%3A25; rur=PRN; urlgen="{\"122.172.28.67\": 24560}:1hyB6m:DZRo2R-Xwv8sSUS3M5WLT2Py4lk"""
    ,
    """mid=XYAJbQALAAHXWOQbCshhjZ0TXF-E; csrftoken=NUYFoEemcuty4uu51cU1Llsn5cpFORht; ds_user_id=20271859208; sessionid=20271859208%3Ad9A8K1aV6VjtYF%3A13; rur=PRN; urlgen="{\"122.172.83.209\": 24560}:1i9zHX:GE2r0_tmukBIbVTdMuLOYr0838w"""""
    ,
    """rur=FTW; mid=XYAK1wALAAHq9WJUOXERpd_OiKcf; csrftoken=rdm6AP9ClCypBcVYSiFckOpVDbekC95E; ds_user_id=20230623790; sessionid=20230623790%3APS7jtLTukDVfWf%3A18; urlgen="{\"122.172.83.209\": 24560}:1i9zNe:ZszM8Y7UTQv-1l-vljNoNiQOeXQ"""""
    ,
    """rur=FTW; mid=XYALWQALAAF0SckkVprBKCYk8gNW; csrftoken=EFb5cgtfSEhOmNFPHhob8fm8VoFGQ2dE; ds_user_id=20477224439; sessionid=20477224439%3AOlbDPDcTuDZEre%3A16; urlgen="{\"122.172.83.209\": 24560}:1i9zPd:CoqGYmV9n0Okid3Vqcg5WxEDCwM"""""
    ,
    """mid=XYAL4gALAAFh_FVF9x4v0OI8xYkP; csrftoken=7illVCwbZFQDG6m3900qh6aqTcBeasRH; ds_user_id=20476600456; sessionid=20476600456%3AptmVsaeke68NrY%3A26; rur=ATN; urlgen="{\"122.172.83.209\": 24560}:1i9zRe:tN-0hsbdY9aEMGh2_cL8LU7VCwA"""""
    ,
    """mcd=3; mid=XEGl1wALAAFdN6IIM4vURanDV6aj; fbm_124024574287414=base_domain=.instagram.com; csrftoken=O7fuOjaW3JnDfgoIhh1y70oEFXgDPFfi; ds_user_id=5589583732; sessionid=5589583732%3A6m0unLcYAAUxz1%3A23; shbid=7908; shbts=1564288697.4193509; rur=PRN; urlgen="{\"122.172.65.106\": 24560\054 \"122.172.119.130\": 24560}:1hsRfq:Jhoy_rzw5b5---7luW2dxt59Rko"""""
    ,

    """mid=XVO-6wALAAHhAyfUfW08OcpmZ4Sg; csrftoken=cNK7eqBZn0ANO68uCfsBw2nd8W3crAZC; shbid=18423; shbts=1565769477.7688148; ds_user_id=6006392694; sessionid=6006392694%3A6IHyf2QSUpYDna%3A29; rur=FTW; urlgen="{\"122.172.28.67\": 24560}:1hxoAS:VMXsdGLoo5tCSXmJQgNGpz-nPRo"""
    ,

    """mid=XVUVmgALAAEadNXlPmfkf1ialoBx; csrftoken=I1na7eBqyTIfttfWjo0DO0r905g4MAav; ds_user_id=18096204756; sessionid=18096204756%3AlgU0EpARkioKHv%3A3; rur=ATN; urlgen="{\"122.172.28.67\": 24560}:1hyAzN:QuPFiCLAPJmGXozT3H5DOxhy7gA"""
    ,

    """mid=XVUWDQALAAGJtRvqfW-_HK1bZ6GH; csrftoken=IIvueDhXQY19Q4KAKOkzcAY4mF65bTMF; ds_user_id=18093461285; sessionid=18093461285%3A7MNLcS0mKg0VRQ%3A22; rur=FTW; urlgen="{\"122.172.28.67\": 24560}:1hyB18:QYrb25b45YmglYiKgVqmoRUkUWs"""
    ,

    """rur=FRC; mid=XVUW7wALAAGqP7EKAhsEaLfUMwDI; csrftoken=0DAnwfGxL1TZEF1a5QgDU5DbCuZNiZAj; ds_user_id=11835832380; sessionid=11835832380%3AX2z0zAjOlrKTf9%3A20; urlgen="{\"122.172.28.67\": 24560}:1hyB4o:4AVNvEEIPoTTRO_4jkAu8S6DRWM"""
    ,

    """mid=XVUXTwALAAE_gRSiZ7oReErEEsM-; csrftoken=NVdeQTMMpNZjD9c5fsBkmFmRwjwJNZkv; ds_user_id=18070511049; sessionid=18070511049%3AxUebJzsAmvkqC1%3A25; rur=PRN; urlgen="{\"122.172.28.67\": 24560}:1hyB6m:DZRo2R-Xwv8sSUS3M5WLT2Py4lk"""

]

chosen_cookie = ''


def get_user(user_id, user_info):
    randInt = random.randint(0, 4)
    try:
        user_url = "https://i.instagram.com/api/v1/users/" + user_id + "/info/"
        response = requests.get(user_url, headers={"cookie": cookie_value[randInt], 'User-Agent': user_agent},
                                timeout=10)

        print(cookie_value[randInt])
        user_data = json.loads(response.text)
        username = user_data['user']['username']
        follower_count = user_data['user']['follower_count']
        try:
            public_email = user_data['user']['public_email']
        except:
            public_email = ' '
        full_name = user_data['user']['full_name']

        user_info.extend([username, follower_count, public_email, full_name])
        return user_info
        # print(
        #     "ID: " + user_id + " " + "Username : " + username + " " + str(score))

    except Exception as e:
        print(e)
        print("The API has been timed out")
        time.sleep(140)


def get_location(shortcode):
    r = ""
    randInt = random.randint(0, 4)
    try:
        url = "https://www.instagram.com/p/" + shortcode + "/?__a=1"
        try:
            r = requests.get(url, headers={"cookie": cookie_value[randInt]}, timeout=10)
        except Exception as e:
            print(e)
            get_location(shortcode)

        data = json.loads(r.text)
        try:
            location = data['graphql']['shortcode_media']['location']['name']  # get location for a post
        except:
            location = ''  # if location is NULL
    except:
        location = ''

    return location
    # print(location)


def move_to_excel(data, location, tag):
    try:
        data.insert(0, location)

        # split the full name
        name_arr = data[4].split(" ", 1)
        first = name_arr[0]
        last = name_arr[1] if len(name_arr) > 1 else ""
        # remove the fullname element and add fname and lname
        data.pop()
        data.append(first)
        data.append(last)
        print(data)
        save_data.append(data)
        if row_count % 100 == 0:

            print("Storing data in bulk YOLO")
            headers = ['Location', 'Username', 'Followers', 'Email', 'First name', 'Last name']

            if row_count % 100000 == 0 and row_count > 0:
                global counter
                counter += 1
                global workbook_name
                workbook_name = tag + str(counter) + ".xlsx"

            global wb
            if os.path.isfile(workbook_name):
                wb = load_workbook(filename=workbook_name)
                sheet = wb.active
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.append(headers)
                for cell in sheet["1:1"]:
                    cell.font = Font(bold=True)

            for d in save_data:
                sheet.append(d)

            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['D'].width = 30
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20

            wb.save(filename=workbook_name)
            save_data.clear()

    except Exception as e:
        print(e)


# sheet.cell(row=row_num + 1, column=1, value=location)
# for i in range(0, 4):
#     sheet.cell(row=row_num + 1, column=i + 2, value=data[i])


def get_location_id(entry):
    location_id = None
    get_location_id_url = "https://www.instagram.com/web/search/topsearch/?context=blended&query=" + entry + "&rank_token=0.20850940886082237&include_reel=true"
    req = requests.get(get_location_id_url, headers={"cookie": cookie_value[0]}, timeout=10)

    location_data = json.loads(req.text)
    places = location_data['places']
    for place in places:
        location_id = place['place']['location']['pk']
        break
    return location_id


def get_location_name(entry_now):
    with open('zip_code_database.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                pass
            else:
                if entry_now == row["zip"]:
                    return row["primary_city"]
            line_count += 1
    return None


def start_scraping(entry, choice):
    global workbook_name
    workbook_name = entry + ".xlsx"
    global row_count
    row_count = 0
    end_cursor = ''
    location_id = None
    abort = False
    if choice is "location":
        location_id = get_location_id(entry)
    if choice is 'zip':
        location_name = get_location_name(entry)
        if location_name is None:
            abort = True
            print('Zipcode 404')

        if abort is False:
            location_id = get_location_id(location_name)

    if abort is False:
        for page in range(num_of_pages):

            try:
                if page == 0:
                    if choice is "tag":
                        url = "https://www.instagram.com/explore/tags/" + entry + "/?__a=1"

                    else:
                        url = "https://www.instagram.com/explore/locations/" + location_id + "/?__a=1"

                else:
                    if choice is "tag":
                        url = "https://www.instagram.com/explore/tags/" + entry + "/?__a=1&max_id=" + end_cursor
                    else:
                        url = "https://www.instagram.com/explore/locations/" + location_id + "/?__a=1&max_id=" + end_cursor

                print(url)
                r = requests.get(url, headers={"cookie": cookie_value[0]}, timeout=10)

                if r.status_code != 200:
                    print("No Posts found. Please Stop scraping before starting a new search")
                    continue

                print(r.text)
                data = json.loads(r.text)

                if choice is "tag":
                    edges = data['graphql']['hashtag']['edge_hashtag_to_media']['edges']  # list with posts
                else:
                    edges = data['graphql']['location']['edge_location_to_media']['edges']  # list with posts

                for item in edges:
                    if stop_thread is True:
                        return
                    while pause_thread:
                        pass

                    try:
                        start_time1 = time.time()
                        post = (item['node'])
                        owner = post['owner']
                        user_id = owner['id']
                        shortcode = post['shortcode']
                        location = get_location(shortcode)

                        user_info = []
                        info = get_user(user_id, user_info)

                        print("--- %s seconds | User Time ---" % round(time.time() - start_time1, 2))
                        start_time2 = time.time()
                        if len(info) != 0:
                            move_to_excel(info, location, entry)
                            row_count += 1
                            print(row_count)

                        print("--- %s seconds --- | Excel time" % round(time.time() - start_time2, 2))

                    except Exception as e:
                        print(e)

                if choice is "tag":
                    end_cursor = data['graphql']['hashtag']['edge_hashtag_to_media']['page_info'][
                        'end_cursor']  # value for the next page
                else:
                    end_cursor = data['graphql']['location']['edge_location_to_media']['page_info'][
                        'end_cursor']  # value for the next page
                if end_cursor is None or end_cursor == "":
                    stop_scraping()
                    return

            except Exception as e:
                print(e)


def get_location_list(entry, choice):
    location = entry
    if choice == "zip":
        location = get_location_name(entry)
        if location is None:
            return None

    get_location_id_url = "https://www.instagram.com/web/search/topsearch/?context=blended&query=" + location + "&rank_token=0.20850940886082237&include_reel=true"
    req = requests.get(get_location_id_url, headers={"cookie": cookie_value[0]}, timeout=10)

    location_data = json.loads(req.text)
    places = location_data['places']

    place_count = 1
    location_list = []
    for place in places:
        location_name = place['place']['location']['name']
        location_list.append(location_name)
        place_count += 1

    return location_list


def stop_scraping():
    global stop_thread
    stop_thread = True

    # deleting the location file - location.txt
    if os.path.isfile('entry.txt'):
        os.remove("entry.txt")

    try:

        headers = ['Location', 'Username', 'Followers', 'Email', 'First name', 'Last name']

        global wb
        if os.path.isfile(workbook_name):
            wb = load_workbook(filename=workbook_name)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(headers)
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)

        for d in save_data:
            sheet.append(d)

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20

        wb.save(filename=workbook_name)
        save_data.clear()

        t1.join()
        stop_thread = False
    except:
        print("Save failed")


def pause_scraping():
    global pause_thread
    if pause_thread:
        pause_thread = False
    else:
        pause_thread = True

    return pause_thread
