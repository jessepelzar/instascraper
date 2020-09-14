import csv
import glob
import json
import os
import sys
import random
import threading
import time
from itertools import cycle
import dload
import re
from time import sleep
# import time
import datetime
# pip3 install selenium
# pip3 install chromedriver
# pip3 install webdriver-manager
#  pip3 install pyvirtualdisplay
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.common.keys import Keys  
# from selenium.webdriver.chrome.options import Options 
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.by import By
# from selenium.common.exceptions import TimeoutException

from InstagramAPI import InstagramAPI
from django.http import HttpResponse, Http404
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .radius import *
from .utils import create_text_file

def update_cookie():
    global cookie, proxy, accounts, proxies, api
    account = next(accounts)
    # proxy = next(proxies)
    proxy = {}
    username, password = account.split(':')
    api = InstagramAPI(username, password)
    # api.setProxy(proxy)
    api.login()
    cookie = '; '.join([f'{key}={value}' for key, value in api.s.cookies.get_dict().items()])


accounts = [
    # 'wilfreddewald2:76859oti',
    # 'jarrettcossette:76859oti',
    # 'rolandwetherington7:76859oti',
    # 'jacquesgoebel:76859oti',
    # 'lucioamyx:76859oti',
    # 'garfieldmoreman:76859oti',
    # 'kellyluedtke3:76859oti',
    # 'hassanlux3:76859oti',
    # 'conniemetoyer:76859oti',
    # 'stacytopping988:76859oti',
    # 'dillonbussiere262:76859oti',
    # 'rickmanganaro3:76859oti',
    # 'haydenbreland4:76859oti',
    # 'shirleyhenegar8:76859oti',
    # 'rodrigocalbert:76859oti',
    # 'octaviorice8:76859oti',
    # 'jeremiahcude3:76859oti',
    # 'staceycollins21811:76859oti',
    # 'gilbertoparish3:76859oti'
    
    
    'santiagozuk99:76859oti',
    'timmycastanon38:76859oti',
    'chasedyer91:76859oti',
    'VitoMaurin:76859oti',
    'parissedberry356:76859oti',
    'bryceblain24:76859oti',
    'burtonkessel3:76859oti',
    'mohammedvillalvazo:76859oti',
    'orvalguillaume:76859oti',
    'reynaldohenline:76859oti',
]

PROXIES = [
    # 'p.webshare.io:20012',
    # 'p.webshare.io:20013',
    # 'p.webshare.io:20014',
    # 'p.webshare.io:20015',
    'p.webshare.io:20000',
    'p.webshare.io:20001',
    'p.webshare.io:20002',
    'p.webshare.io:20003',
    'p.webshare.io:20004',
    'p.webshare.io:20005',
    'p.webshare.io:20006',
    'p.webshare.io:20007',
    'p.webshare.io:20008',
    'p.webshare.io:20009',
    'p.webshare.io:20010',
    'p.webshare.io:20011',
    'p.webshare.io:20012',
    'p.webshare.io:20013',
    'p.webshare.io:20014',
    'p.webshare.io:20015',
    'p.webshare.io:20016',
    'p.webshare.io:20017',
    'p.webshare.io:20018',
    'p.webshare.io:20019',
    'p.webshare.io:20020',
    'p.webshare.io:20021',
    'p.webshare.io:20022',
    'p.webshare.io:20023',
    'p.webshare.io:20024',
]
accounts = cycle(accounts)
PROXIES = cycle(PROXIES)
PROXY = next(PROXIES)
# Create your views here.

def faq(request):
    return render(request, 'scraper/faq.html')


def radius_check(request):
    if request.method == 'POST':
        city_name_r = request.POST.get('radius_value')
        city_list, dist_list = get_cities(city_name_r)
        city_list = zip(city_list, dist_list)
        context = {
            'city_list': city_list,
            'city_name_r': city_name_r,
        }
        return render(request, 'scraper/radius.html', context)

    return render(request, 'scraper/radius.html')


def show(request):
    if request.method == 'POST':
        file_name_r = request.POST.get('filename')
        file_path = os.path.join(file_name_r)
        if os.path.exists(file_path):

            if request.POST.get('delete_file'):
                # deleting the file
                os.remove(file_path)
                file_names = []
                for file in glob.glob("*.xlsx"):
                    file_names.append(file)
                return render(request, 'scraper/show.html', {"file_names": file_names})
            else:
                # downloading the file
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + json.dumps(os.path.basename(file_path))
                    return response

        raise Http404

    else:
        file_names = []
        for file in glob.glob("*.xlsx"):
            file_names.append(file)
        return render(request, 'scraper/show.html', {"file_names": file_names})


@csrf_exempt
def row_ajax(request):
    if request.method == 'POST':
        return HttpResponse(row_count)


pause_thread = False


def stop_scrap(request):
    if request.method == 'POST':
        if request.POST.get('stop_scrap'):
            
            stop_scraping()
            # for thread in thread_list:
            #     if thread.is_alive():
            #         print("thread still alive man. Fuck")
            return render(request, 'scraper/index.html')
        else:
            global pause_value
            pause_value = pause_scraping()  # stop the scraper

            if os.path.isfile('entry.txt'):
                f = open("entry.txt", "r")
                contents = f.read()
                context = {
                    "row_count": row_count,
                    "entry": contents,
                    "running": "True",
                    "pause_scrap": pause_value,
                }
                return render(request, 'scraper/index.html', context)

        return render(request, 'scraper/index.html')

thread_list = []
def index(request):
    if request.method == 'POST':

        global entry_r
        global choice_r

        entry_r = []
        # --------------
        multiple = False
        # --------------
        hashtag_r = request.POST.get('hashtag')
        location_r = request.POST.get('location')
        zip_r = request.POST.get('zip')
        filename_r = request.POST.get('filename')
        hashtag_list_r = request.POST.get('hashtag-list')
        tag_num_switch_r = request.POST.get('tagwithnumberswitch')
        # print(tag_num_switch_r)
        hashtag_list_r = str(hashtag_list_r)
        hashtag_list_r = hashtag_list_r.split(',')
        print(hashtag_list_r)
        
        print("number of hashtags:", len(hashtag_list_r))
        
        if len(hashtag_list_r) != 0:
            choice_r = "tag"
            # entry_r = hashtag_r
            entry_r.clear()
            entry_r = hashtag_list_r
        elif zip_r != "":
            choice_r = "zip"
            entry_r.clear()
            entry_r.append(zip_r)
        elif location_r != "":
            choice_r = "location"
            entry_r.clear()
            entry_r.append(location_r)
        else:
            stop_scraping()
        if request.POST.get('startscraping'):
            global row_count, api_call_count
            row_count = 0
            api_call_count = 0

            update_cookie()
            create_text_file(filename_r)

            cookie_idx = 0
            thread_idx = 0
            for entry in entry_r:
                print("entry", entry)   
                stop_thread.append(False)             
                thread = threading.Thread(target=start_scraping, args=(entry, choice_r, filename_r, tag_num_switch_r, cookie_idx, thread_idx))
                thread_list.append(thread)
                cookie_idx += 1 # just keeping separate for cookie and thread for now
                thread_idx += 1
            
            for thread in thread_list:
                # thread.daemon = True
                thread.start()

            if multiple is True:
                if len(entry_r) > 0:
                    # print(row_count)
                    context = {
                        "row_count": row_count,
                        "entry": entry_r[0],
                        "running": "True",
                    }
            else:
                if len(entry_r) > 0:
                    # print(row_count)
                    context = {
                        "row_count": row_count,
                        "entry": filename_r,
                        "running": "True",
                    }
                return render(request, 'scraper/index.html', context)

        elif request.POST.get('checklocation'):
            if multiple is True:
                location_list = get_location_list(entry_r[0], choice_r)
                context = {
                    "location_list": location_list,
                    "entry": filename_r,
                }
            else:
                location_list = get_location_list(entry_r[0], choice_r)
                context = {
                    "location_list": location_list,
                    "entry": filename_r,
                }
                return render(request, 'scraper/index.html', context)

        #  elif request.POST.get('checklocation'):
        #     if choice_r == "tagAndLocation":
        #         location_list = get_location_and_tag_list(entry_r[0], entry_r[1], choice_r)
        #         context = {
        #             "location_list": location_list,
        #             "entry": entry_r,
        #         }
        #         return render(request, 'scraper/index.html', context)

        else:
            pass

    else:
        if os.path.isfile('entry.txt'):
            f = open("entry.txt", "r")
            contents = f.read()
            context = {
                "row_count": row_count,
                "entry": contents,
                "running": "True",
                "pause_scrap": pause_thread,
            }
            return render(request, 'scraper/index.html', context)
        else:

            return render(request, 'scraper/index.html')


num_of_pages = 500000
row_count = 0
stop_thread = []

save_data = []
user_agent = "Mozilla/5.0 (iPhone; CPU iPhone OS 12_3_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 Instagram 105.0.0.11.118 (iPhone11,8; iOS 12_3_1; en_US; en-US; scale=2.00; 828x1792; 165586599)"


def get_user(user_id, user_info):
    global PROXY, PROXIES
    global api
    print("---------- get user ----------")
    api.getUsernameInfo(user_id)
    if api.LastResponse.status_code == 429:
        sleep(86400)
        update_cookie()
        return user_info
    elif api.LastResponse.status_code == 400:
        stop_scraping()
    user_data = json.loads(api.LastResponse.text)
    # print(user_data)
    username = user_data['user']['username']
    follower_count = user_data['user']['follower_count']
    try:
        public_email = user_data['user']['public_email']
    except:
        public_email = ' '
    full_name = user_data['user']['full_name']
    igURL = 'https://www.instagram.com/' + username + '/'

    # user_info.extend([username, userFirstName, userLastName, public_email, followers, following, external_url, numberOfPosts, igURL])
    user_info.extend([username, full_name, public_email, follower_count, igURL])
    return user_info, username
        # print(
        #     "ID: " + user_id + " " + "Username : " + username + " " + str(score))



def start_scraping(entry, choice, filename_r, tag_num_switch_r, cookie_idx, thread_idx):

    # COOKIES = cycle(cookie_value[cookie_idx])
    print("---------- start scraping ----------")
    # print(choice)
    # print("switch", tag_num_switch_r)
    global workbook_name, cookie, proxy
    workbook_name = filename_r + ".xlsx"
    # if choice is 'tagAndLocation':
    #     workbook_name = entry[0] + "_" + entry[1] + ".xlsx"
    # else:
    #     workbook_name = entry + ".xlsx"
    global row_count, api_call_count
    row_count = 0
    end_cursor = ''
    location_id = None
    # entryChosen = entry
    abort = False
    # if choice is 'tagAndLocation':
    #     print("tag and location chosen")
    if choice is 'tag':
        print("tag chosen")
    if choice is "location":
        location_id = get_location_id(entry)
        print("location id", location_id)
    if choice is 'zip':
        location_name = get_location_name(entry)
        if location_name is None:
            abort = True
            print('Zipcode 404')
        if abort is False:
            location_id = get_location_id(location_name)

    # sys.exit()
    if abort is False:
        dateCounter = 0
        for page in range(num_of_pages):
            print("---------- pages loop ----------")
            # print(COOKIE)
            entryChosen = None
            try:
                print("---------- try get page ----------")
                # if entry == "" and location_id == None:
                #     print("scraping stopped")
                #     stop_scraping()

                if page == 0:
                    if choice is "tag":
                        entryChosen = entry.replace(" ", "")
                        url = "https://www.instagram.com/explore/tags/" + entryChosen + "/?__a=1"

                    else:
                        url = "https://www.instagram.com/explore/locations/" + location_id + "/?__a=1"

                else:
                    if choice is "tag":
                        entryChosen = entry.replace(" ", "")
                        url = "https://www.instagram.com/explore/tags/" + entryChosen + "/?__a=1&max_id=" + end_cursor
                    else:
                        url = "https://www.instagram.com/explore/locations/" + location_id + "/?__a=1&max_id=" + end_cursor

                while True:
                    r = requests.get(url, headers={"cookie": cookie, "User-Agent": user_agent}, timeout=60, proxies={'http': f'http:{PROXY}', 'https': f'https:{PROXY}'})
                    
                    if r.status_code == 200:
                        print("code is 200 fuck")
                        break
                    elif r.status_code in [400, 429]:
                        print("cookie update")
                        sleep(random.randint(30, 60))
                        update_cookie()

                print(f'>>>>>>>>>>>>>> {r.status_code} <<<<<<<<<<<<<<<')
                data = json.loads(r.text)
                
                if choice is "tag":
                    edges = data['graphql']['hashtag']['edge_hashtag_to_media']['edges']  # list with posts
                else:
                    edges = data['graphql']['location']['edge_location_to_media']['edges']  # list with posts

                for item in edges:
                    if stop_thread[thread_idx] is True:
                        return
                    while pause_thread:
                        pass

                    try:
                        start_time1 = time.time()
                        post = (item['node'])
                        owner = post['owner']
                        user_id = owner['id']
                        shortcode = post['shortcode']
                        location, country_code = get_location(shortcode)
                        if location == "": 
                            print("skip")
                            sleep(random.randint(2,3))
                            continue

                        if country_code != "US": 
                            print("not us")
                            sleep(random.randint(2,3))
                            continue

                        timestamp = post['taken_at_timestamp']
                        user_info = []
                        sleep(random.randint(5,6))
                        print("before get user")
                        info, username = get_user(user_id, user_info)
                        print("after get user")
                        if choice is "tag":
                            if str(tag_num_switch_r) == "true":
                                future_date = get_future_date(timestamp, entryChosen)
                                if future_date is None:
                                    dateCounter += 1
                                    print(thread_list)
                                    if dateCounter > 5:
                                        print("-----------------------------------")
                                        dateCounter = 0
                                        kill_single_thread(thread_idx)
                                        abort = True
                                        return
                                else:
                                    dateCounter = 0
                                # print("date counter", dateCounter)
                                info.extend([future_date, entryChosen])
                                
                        print(info)
                        print(location)
                        print("--- %s seconds | User Time ---" % round(time.time() - start_time1, 2))
                        start_time2 = time.time()
                        if len(info) != 0:
                            move_to_excel(info, location, entryChosen)
                            row_count += 1
                            api_call_count += 1
                            if api_call_count >= random.randint(150, 250):
                                api_call_count = 0
                                print('>>>>>>>>> Updating Cookie <<<<<<<<<')
                                sleep(random.randint(30, 60))
                                update_cookie()
                            print(row_count)

                        print("--- %s seconds --- | Excel time" % round(time.time() - start_time2, 2))

                    except Exception as e:
                        print(e)
                        # sys.exit()
                        

                if choice is "tag":
                    end_cursor = data['graphql']['hashtag']['edge_hashtag_to_media']['page_info']['end_cursor']  # value for the next page
                else:
                    end_cursor = data['graphql']['location']['edge_location_to_media']['page_info']['end_cursor']  # value for the next page
                if end_cursor is None or end_cursor == "":
                    stop_scraping()
                    return

            except Exception as e:
                # sys.exit()
                print("---------- pages failed ----------")
                print(e)
                # stop_scraping()

def kill_single_thread(thread_idx):
    global stop_thread
    print("thread idx", thread_idx)
    if len(thread_list) - thread_list.count(None) == 1 or len(thread_list) == 1:
        print("last thread")
        stop_scraping()
    else:
        print("stop single thread")
        stop_thread[thread_idx] = True
        sleep(1)
        thread_list[thread_idx].join()
        thread_list[thread_idx] = None
        sleep(1)
        stop_thread[thread_idx] = False
     

def get_future_date(timestamp, tagwithnumber):

    print("get future date")
    # user_url_data = "https://www.instagram.com/p/" + shortcode + "/?__a=1"
    daysTotalPregnant = 280

    numberStr = ""
    for i in range(len(tagwithnumber)):
        if tagwithnumber[i].isdigit():
            numberStr += tagwithnumber[i]
    if numberStr == "":
        return None
        
    tagWeek = int(numberStr)
    tagDays = tagWeek * 7
    # switch_count = 0
    # while switch_count < 5:
    # print(f'SWITCH COUNT SWITCH COUNT {switch_count}')
    # try: 
    #     data_response = requests.get(user_url_data, headers={"cookie": COOKIE, 'User-Agent': user_agent}, timeout=10)
    #     # break
    # except:
    #     PROXY = next(PROXIES)
    #     switch_count+=1
    # if switch_count == 5:
    #     return None
        # print(cookie)

    # data = json.loads(data_response.text)
    # timestamp = data['graphql']['shortcode_media']['taken_at_timestamp']

    
    

    postDate = datetime.datetime.fromtimestamp(timestamp)
    postDateList = [postDate.month, postDate.day, postDate.year]
    postDayOfYear = dayOfYear(postDateList[0], postDateList[1], postDateList[2])

    todaysDate = datetime.datetime.today()
    todayDayList = [todaysDate.month, todaysDate.day, todaysDate.year]
    todayDayOfYear = dayOfYear(todayDayList[0], todayDayList[1], todayDayList[2])

    daysSincePost = todayDayOfYear - postDayOfYear
    daysPreg = daysSincePost + tagDays
    daysLeft = daysTotalPregnant - daysPreg
    formatedDaysLeft = datetime.timedelta(days=daysLeft)

    dueDate = todaysDate + formatedDaysLeft
    dueDayOfYear = dayOfYear


    daysRemaining = (dueDate - todaysDate).days
    if daysRemaining < 0:
        return None
    # projectedDueDayOfYear = todayDayOfYear + daysLeft
    # print(todayDayOfYear)
    # print(postDayOfYear)
    # print(daysPreg)
    # print(daysLeft)
    # print(dueDate.date())
    # print('day of year', todayDayOfYear)
    
    
    # user_info.extend([dueDate])
    return str(dueDate.date())
    
    # -------------------------
    # -------------------------

def dayOfYear(month, day, year):
    # date = '%s-%s-%s' % (year, month, day)
    days = [0,31,28,31,30,31,30,31,31,30,31,30,31]
    # d = list(map(int, date.split("-")))
    d = [year, month, day]
    if d[0] % 400 == 0:
        days[2]+=1
    elif d[0]%4 == 0 and d[0]%100!=0:
        days[2]+=1
    for i in range(1,len(days)):
        days[i]+=days[i-1]
    return days[d[1]-1]+d[2]

def get_location(shortcode):
    global cookie, proxy
    country_code = ""
    r = ""
    try:
        url = "https://www.instagram.com/p/" + shortcode + "/?__a=1"
        try:
            r = requests.get(url, headers={"cookie": cookie}, timeout=60)
        except Exception as e:
            print(e)
            # if ("char 0" in e):
            print("fuck1")
            get_location(shortcode)
            stop_scraping()

        data = json.loads(r.text)
        # print(data)
        try:
            location = data['graphql']['shortcode_media']['location']['name']  # get location for a post
            location2 = data['graphql']['shortcode_media']['location']['address_json']
            location3 = json.loads(location2)
            country_code = location3['country_code']
            print(f'>>>>>>>>>>>>>>{country_code}<<<<<<<<<<<<<')
        except:
            location = ''  # if location is NULL
    except:
        location = ''

    return location, country_code


def move_to_excel(data, location, tag):
    try:
        data.insert(0, location)
        save_data.append(data)
        if row_count % 100 == 0:

            # print("Storing data in bulk YOLO")
            # headers = ['Location','Username','First Name', 'Last Name', 'Public Email', 'Followers', 'Following', 'External URL', 'Number of Posts', 'Profile URL', 'Due Date', 'Tag']
            headers = ['Location','Username','Full Name', 'Public Email', 'Followers', 'Profile URL', 'Due Date', 'Tag']
            
            if row_count % 100000 == 0 and row_count > 0:
                global counter
                counter += 1
                global workbook_name
                workbook_name = tag + str(counter) + ".xlsx"

            global wb
            if os.path.isfile(workbook_name):
                wb = load_workbook(filename=workbook_name)
                sheet = wb.active
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.append(headers)
                for cell in sheet["1:1"]:
                    cell.font = Font(bold=True)

            for d in save_data:
                sheet.append(d)

            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 30
            sheet.column_dimensions['F'].width = 10
            sheet.column_dimensions['G'].width = 10
            sheet.column_dimensions['H'].width = 30
            sheet.column_dimensions['I'].width = 10
            sheet.column_dimensions['J'].width = 30
            sheet.column_dimensions['K'].width = 30
            sheet.column_dimensions['L'].width = 30

            wb.save(filename=workbook_name)
            save_data.clear()

    except Exception as e:
        print(e)


# sheet.cell(row=row_num + 1, column=1, value=location)
# for i in range(0, 4):
#     sheet.cell(row=row_num + 1, column=i + 2, value=data[i])


def get_location_id(entry):
    global cookie, proxy
    location_id = None
    get_location_id_url = "https://www.instagram.com/web/search/topsearch/?context=blended&query=" + entry + "&rank_token=0.20850940886082237&include_reel=true"
    req = requests.get(get_location_id_url, headers={"cookie": cookie}, timeout=60)

    location_data = json.loads(req.text)
    places = location_data['places']
    for place in places:
        location_id = place['place']['location']['pk']
        break
    return location_id


def get_location_name(entry_now):
    with open('zip_code_database.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                pass
            else:
                if entry_now == row["zip"]:
                    return row["primary_city"]
            line_count += 1
    return None

def get_location_list(entry, choice):
    global cookie, proxy
    location = entry
    if choice == "zip":
        location = get_location_name(entry)
        if location is None:
            return None

    get_location_id_url = "https://www.instagram.com/web/search/topsearch/?context=blended&query=" + location + "&rank_token=0.20850940886082237&include_reel=true"
    req = requests.get(get_location_id_url, headers={"cookie": cookie}, timeout=60)

    location_data = json.loads(req.text)
    places = location_data['places']
    place_count = 1
    location_list = []
    for place in places:
        location_name = place['place']['location']['name']
        location_list.append(location_name)
        place_count += 1

    return location_list


def stop_scraping():
    
    global stop_thread
    # stop_thread = True
    stop_thread_len = len(stop_thread)
    stop_thread = [True] * stop_thread_len
    for thread in thread_list:
        if thread is None:
            continue
        else:
            thread.join()
    thread_list.clear()
    # deleting the location file - location.txt
    if os.path.isfile('entry.txt'):
        os.remove("entry.txt")

    try:

        # headers = ['Location','Username','First Name', 'Last Name', 'Public Email', 'Followers', 'Following', 'External URL', 'Number of Posts', 'Profile URL', 'Due Date', 'Tag']
        headers = ['Location','Username','Full Name', 'Public Email', 'Followers', 'Profile URL', 'Due Date', 'Tag']
        global wb
        if os.path.isfile(workbook_name):
            wb = load_workbook(filename=workbook_name)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(headers)
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)

        for d in save_data:
            sheet.append(d)

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 30
        sheet.column_dimensions['L'].width = 30

        wb.save(filename=workbook_name)
        save_data.clear()
        # stop_thread = False
        stop_thread = [False] * stop_thread_len
    except:
        print("Save failed")


def pause_scraping():
    global pause_thread
    if pause_thread:
        pause_thread = False
    else:
        pause_thread = True

    return pause_thread
